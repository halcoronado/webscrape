
from tkinter import W
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font





#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2022/'

page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

title = soup.title

print(title.text)

movie_table = soup.find('table')
table_rows = movie_table.findAll("tr")
#print(table_rows[1])

##
##
##
##




### movie number 1-5, name, gross, total gross

for x in range(1,6):
    td = table_rows[x].findAll('td')
    ranking = td[0].text
    title =td[1].text
    gross=td[5].text
    total_gross=td[7].text

    print(gross)

wb= xl.Workbook()

ws =wb.active

ws.title = 'Box Office Report'


fontobject = Font (size=15, bold= True)
ws['A1']= 'No.'
ws['A1'].font = fontobject
ws['B1']= 'Movie Title'
ws['B1'].font = fontobject
ws['C1']= 'Release Date'
ws['C1'].font = fontobject
ws['D1']= 'Gross'
ws['D1'].font = fontobject
ws['E1']= 'Total Gross'
ws['E1'].font = fontobject
ws['F1']= "% of Total Gross"
ws['F1'].font = fontobject


for x in range(1,6):
    td = table_rows[x].findAll('td')
    ranking = td[0].text
    title =td[1].text
    gross=int(td[5].text.replace(',','').replace("$",''))
    total_gross=int(td[7].text.replace(',','').replace("$",''))
    release_date = td[8].text

    percent_gross = round((gross/total_gross)*100,2)
    print(gross)
    ws['A'+str(x+1)]=ranking
    ws['B'+str(x+1)]=title
    ws['C'+str(x+1)]= release_date
    ws['D'+str(x+1)]= gross
    ws['E'+str(x+1)]=total_gross
    ws['F'+str(x+1)]= str(percent_gross)+'%'


ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 26
##for cell in ws[1:1]:
    #cell.font =fontobject


wb.save('Top5_GrossingMovies.xlsx')
